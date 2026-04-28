from __future__ import annotations

from collections import Counter
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest

from app.modules.budget_constructor import (
    compare_dataset,
    load_task_dataset,
    query_dataset,
    query_result_to_csv,
    query_result_to_xlsx,
    search_dataset,
    timeline_dataset,
)
from app.modules.budget_constructor.parsing import kcsr_slice
from app.modules.budget_constructor.types import (
    SOURCE_AGREEMENTS,
    SOURCE_BUAU,
    SOURCE_GZ_BUDGET_LINES,
    SOURCE_GZ_CONTRACTS,
    SOURCE_GZ_PAYMENTS,
    SOURCE_RCHB,
)


REPO_ROOT = Path(__file__).resolve().parents[3]
TASK_DIR = REPO_ROOT / "task"

pytestmark = pytest.mark.skipif(not TASK_DIR.exists(), reason="task data folder is not available")


@pytest.fixture(scope="session")
def dataset():
    return load_task_dataset(TASK_DIR)


def test_use_case_admin_imports_real_task_sources_and_gets_quality_profile(dataset):
    source_counts = Counter(source.source_type for source in dataset.source_files)

    assert source_counts[SOURCE_RCHB] == 15
    assert source_counts[SOURCE_AGREEMENTS] >= 15
    assert source_counts[SOURCE_GZ_BUDGET_LINES] == 1
    assert source_counts[SOURCE_GZ_CONTRACTS] == 1
    assert source_counts[SOURCE_GZ_PAYMENTS] == 1
    assert source_counts[SOURCE_BUAU] == 5

    assert len(dataset.budget_facts) > 1000
    assert len(dataset.agreements) > 1000
    assert len(dataset.contracts) == 22
    assert len(dataset.contract_budget_lines) == 31
    assert len(dataset.payments) == 323

    issue_codes = Counter(issue.code for issue in dataset.issues)
    assert issue_codes["payment_contract_missing"] > 0
    assert issue_codes["missing_columns"] == 0
    assert any(source.rows_imported > 0 for source in dataset.source_files)
    assert all(
        source.rows_imported > 0
        for source in dataset.source_files
        if source.source_type == SOURCE_RCHB and source.period_date and source.period_date.year == 2026
    )


def test_use_case_user_searches_by_code_city_and_contract_number(dataset):
    code_hits = search_dataset(dataset, "6105")
    city_hits = search_dataset(dataset, "Тында")
    contract_hits = search_dataset(dataset, "Ф.2025")

    assert code_hits
    assert any(hit.matched_codes.get("kcsr") and "6105" in hit.matched_codes["kcsr"] for hit in code_hits)
    assert city_hits
    assert any(SOURCE_RCHB in hit.source_types for hit in city_hits)
    assert contract_hits


def test_use_case_user_builds_skk_template_selection(dataset):
    result = query_dataset(
        dataset,
        template_code="skk",
        date_from=date(2025, 1, 1),
        date_to=date(2026, 1, 1),
        metrics=["LIMITS", "BO", "CASH_RCHB", "AGREEMENT_MBT", "CONTRACT_AMOUNT", "CONTRACT_PAYMENT"],
    )

    assert result.rows
    assert result.totals["LIMITS"] > 0
    assert result.totals["BO"] > 0
    assert result.totals["CASH_RCHB"] > 0
    assert result.totals["AGREEMENT_MBT"] > 0
    assert any(row.drilldown_available for row in result.rows)
    assert all(kcsr_slice(row.codes.get("kcsr"), 6, 4) == "6105" for row in result.rows if row.codes.get("kcsr"))
    assert all("payment_contract_missing" != warning.code for warning in result.warnings)


def test_use_case_query_warnings_are_limited_to_result_rows(dataset):
    result = query_dataset(
        dataset,
        template_code="skk",
        date_from=date(2025, 1, 1),
        date_to=date(2026, 4, 1),
        metrics=["CONTRACT_AMOUNT", "CONTRACT_PAYMENT"],
    )

    assert result.rows
    warning_codes = {warning.code for warning in result.warnings}
    assert warning_codes <= {"contract_amount_allocated_equally"}
    contract_ids = {
        record.details.get("con_document_id")
        for records in result.drilldowns.values()
        for record in records
        if record.details.get("con_document_id")
    }
    assert all(
        any(contract_id in warning.message for contract_id in contract_ids)
        for warning in result.warnings
    )


def test_use_case_control_templates_are_executable_and_real_data_is_explicit(dataset):
    metrics = ["AGREEMENT_MBT", "AGREEMENT_SUBSIDY", "CONTRACT_AMOUNT", "BUAU_PAYMENT"]

    for template_code in ("skk", "two_three", "okv"):
        result = query_dataset(
            dataset,
            template_code=template_code,
            date_from=date(2025, 1, 1),
            date_to=date(2026, 4, 1),
            metrics=metrics,
        )
        assert result.rows, template_code
        assert sum(result.totals.values(), Decimal("0.00")) > 0

    kik_result = query_dataset(
        dataset,
        template_code="kik",
        date_from=date(2025, 1, 1),
        date_to=date(2026, 4, 1),
        metrics=metrics,
    )
    assert kik_result.rows == []


def test_use_case_user_compares_periods(dataset):
    result = compare_dataset(
        dataset,
        template_code="skk",
        base_date=date(2025, 2, 1),
        compare_date=date(2026, 1, 1),
        metrics=["LIMITS", "BO", "CASH_RCHB"],
    )

    assert result.rows
    assert any(row.delta != 0 for row in result.rows)
    for row in result.rows:
        assert row.delta == row.compare_value - row.base_value


def test_use_case_user_sees_monthly_timeline(dataset):
    points = timeline_dataset(
        dataset,
        template_code="skk",
        date_from=date(2025, 1, 1),
        date_to=date(2026, 4, 1),
        metrics=["LIMITS", "BO", "CASH_RCHB"],
    )

    assert points
    assert [point.period for point in points] == sorted(point.period for point in points)
    assert any(point.metric_code == "CASH_RCHB" and point.amount > 0 for point in points)


def test_use_case_user_opens_drilldown_and_exports_selection(dataset, tmp_path):
    result = query_dataset(
        dataset,
        template_code="skk",
        date_from=date(2025, 1, 1),
        date_to=date(2026, 4, 1),
        metrics=["AGREEMENT_MBT"],
    )
    row = next(row for row in result.rows if row.drilldown_available)
    drilldown = result.drilldowns[row.row_id]

    assert drilldown
    assert any(record.details.get("reg_number") for record in drilldown)
    assert any(record.details.get("recipient") for record in drilldown)

    csv_text = query_result_to_csv(result)
    assert "Объект,Показатель,Сумма" in csv_text
    assert any(item.object_name.split('"')[0] in csv_text for item in result.rows)

    xlsx_path = query_result_to_xlsx(result, tmp_path / "selection.xlsx")
    assert xlsx_path.exists()

    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, read_only=True)
    assert "Выборка" in workbook.sheetnames
