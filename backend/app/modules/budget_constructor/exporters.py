from __future__ import annotations

import csv
from io import StringIO
from pathlib import Path

from app.modules.budget_constructor.types import QueryResult


CSV_COLUMNS = [
    ("object_name", "Объект"),
    ("metric_name", "Показатель"),
    ("amount", "Сумма"),
    ("source_type", "Источник"),
    ("warning_codes", "Предупреждения"),
]

SOURCE_LABELS = {
    "rchb": "РЧБ",
    "agreements": "Соглашения",
    "gz_budget_lines": "ГЗ: бюджетные строки",
    "gz_contracts": "ГЗ: договоры и контракты",
    "gz_payments": "ГЗ: платежи",
    "buau": "БУАУ",
}

WARNING_LABELS = {
    "equal_by_line_no_amount": "Сумма распределена поровну",
    "contract_amount_allocated_equally": "Сумма распределена поровну",
    "payment_budget_line_missing": "Платеж без бюджетной строки",
    "payment_contract_missing": "Платеж без договора",
    "contract_budget_line_missing": "Договор без бюджетной строки",
    "missing_columns": "Не хватает колонок",
    "row_parse_error": "Ошибка разбора строки",
    "unknown_template": "Неизвестный шаблон",
}


def query_result_to_csv(result: QueryResult) -> str:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([caption for _, caption in CSV_COLUMNS])
    for row in result.rows:
        writer.writerow(
            [
                row.object_name,
                row.metric_name,
                f"{row.amount:.2f}",
                _source_label(row.source_type),
                ", ".join(_warning_label(code) for code in row.warning_codes),
            ]
        )
    return output.getvalue()


def query_result_to_xlsx(result: QueryResult, path: Path | str) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output_path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Выборка"
    ws.append([caption for _, caption in CSV_COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in result.rows:
        ws.append(
            [
                row.object_name,
                row.metric_name,
                float(row.amount),
                _source_label(row.source_type),
                ", ".join(_warning_label(code) for code in row.warning_codes),
            ]
        )

    for col_idx in range(1, len(CSV_COLUMNS) + 1):
        width = max(len(str(ws.cell(row_idx, col_idx).value or "")) for row_idx in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 70)

    for cell in ws["C"][1:]:
        cell.number_format = '# ##0.00'

    if result.warnings:
        issues_ws = wb.create_sheet("Предупреждения")
        issues_ws.append(["Уровень", "Проверка", "Сообщение"])
        for issue in result.warnings:
            issues_ws.append([_severity_label(issue.severity), _warning_label(issue.code), issue.message])
        for cell in issues_ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for col_idx in range(1, 4):
            width = max(len(str(issues_ws.cell(row_idx, col_idx).value or "")) for row_idx in range(1, issues_ws.max_row + 1))
            issues_ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 100)

    wb.save(output_path)
    return output_path


def _source_label(value: str) -> str:
    return SOURCE_LABELS.get(value, value)


def _warning_label(value: str) -> str:
    return WARNING_LABELS.get(value, value)


def _severity_label(value: str) -> str:
    return {"error": "Ошибка", "warning": "Предупреждение"}.get(value, value)
